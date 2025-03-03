from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.generic import ListView, CreateView, DeleteView
from .models import TableData, Filter
from .forms import TableUploadForm, FilterForm, ReportFilterForm, TableDataAppendForm
import pandas as pd
import json
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

def index(request):
    # Получаем все данные
    all_data = []
    for table in TableData.objects.all():
        all_data.extend(table.data_json)
    
    # Разделяем данные на дебет и кредит
    debit_data = []
    credit_data = []
    
    for record in all_data:
        if record['Дебит'] != 'nan':
            debit_data.append({
                'Валюта': 'RUB' if record.get('Валюта', '') == 'nan' else record.get('Валюта', ''),
                'Дата': record.get('Дата', ''),
                'Дебит': record.get('Дебит', ''),
                'Банковское назначение платежа': record.get('Банковское назначение платежа', ''),
                'Контрагент': record.get('Контрагент', ''),
                'Внутренние назначения': record.get('Внутренние назначения', '')
            })
        if record['Кредит'] != 'nan':
            credit_data.append({
                'Валюта': 'RUB' if record.get('Валюта', '') == 'nan' else record.get('Валюта', ''),
                'Дата': record.get('Дата', ''),
                'Кредит': record.get('Кредит', ''),
                'Банковское назначение платежа': record.get('Банковское назначение платежа', ''),
                'Контрагент': record.get('Контрагент', ''),
                'Внутренние назначения': record.get('Внутренние назначения', '')
            })

    return render(request, 'table_app/index.html', {
        'upload_form': TableUploadForm(),
        'filter_form': FilterForm(),
        'report_form': ReportFilterForm(),
        'debit_data': debit_data,
        'credit_data': credit_data,
    })

def upload_file(request):
    if request.method == 'POST':
        form = TableUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Чтение файла
                file = request.FILES['file']
                file_extension = file.name.split('.')[-1].lower()
                
                if file_extension == 'csv':
                    df = pd.read_csv(file)
                elif file_extension == 'ods':
                    df = pd.read_excel(file, engine='odf')
                else:
                    df = pd.read_excel(file)
                
                # Проверка обязательных колонок
                required_columns = ["Дата", "Дебит", "Кредит", "Банковское назначение платежа"]
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f"В файле отсутствуют обязательные колонки: {', '.join(missing_columns)}")
                    return redirect('index')
                
                # Добавление опциональных колонок
                optional_columns = ["Валюта", "Контрагент", "Внутренние назначения"]
                for col in optional_columns:
                    if col not in df.columns:
                        df[col] = ""
                
                # Проверка данных
                if df.empty:
                    messages.error(request, "Файл не содержит данных")
                    return redirect('index')
                
                # Преобразование в JSON
                json_data = []
                for _, row in df.iterrows():
                    record = {}
                    for col in df.columns:
                        value = row[col]
                        if isinstance(value, pd.Timestamp):
                            value = value.strftime('%Y-%m-%d')
                        elif pd.isna(value):
                            value = "nan"
                        else:
                            value = str(value)
                        record[col] = value
                    json_data.append(record)
                
                # Сохранение данных
                table_data = form.save(commit=False)
                table_data.data_json = json_data
                table_data.save()
                
                messages.success(request, f"Данные успешно сохранены на лист '{form.cleaned_data['sheet_name']}'")
            except pd.errors.EmptyDataError:
                messages.error(request, "Файл пуст или имеет неверный формат")
            except pd.errors.ParserError:
                messages.error(request, "Ошибка при чтении файла. Проверьте формат файла")
            except Exception as e:
                messages.error(request, f"Ошибка при загрузке файла: {str(e)}")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                field_name = form.fields[field].label or field
                error_messages.append(f"{field_name}: {', '.join(errors)}")
            messages.error(request, f"Форма заполнена неверно: {'; '.join(error_messages)}")
    return redirect('index')

def apply_filters(request):
    if request.method == 'POST':
        form = FilterForm(request.POST)
        if form.is_valid():
            try:
                # Получаем значения из формы
                contractor = form.cleaned_data.get('contractor', '').strip().lower()
                payment_purpose = form.cleaned_data.get('payment_purpose', '').strip().lower()
                internal_purpose = form.cleaned_data.get('internal_purpose', '').strip()
                
                # Сохраняем фильтр в базу данных
                filter_obj = form.save()
                
                # Получаем все записи из базы данных
                table_data = TableData.objects.all()
                
                # Для каждой записи в базе данных
                for table in table_data:
                    modified = False
                    data_json = table.data_json
                    
                    for record in data_json:
                        # Получаем значения из записи и приводим к нижнему регистру для сравнения
                        record_contractor = record.get('Контрагент', '').strip().lower()
                        record_purpose = record.get('Банковское назначение платежа', '').strip().lower()
                        
                        # Проверяем условия фильтрации с частичным совпадением
                        contractor_match = not contractor or contractor in record_contractor
                        purpose_match = not payment_purpose or payment_purpose in record_purpose
                        
                        # Применяем логику фильтрации
                        if internal_purpose:
                            # Случай 1: Заполнены первый и третий списки
                            if contractor and not payment_purpose and contractor_match:
                                record['Внутренние назначения'] = internal_purpose
                                modified = True
                            
                            # Случай 2: Заполнены второй и третий списки
                            elif payment_purpose and not contractor and purpose_match:
                                record['Внутренние назначения'] = internal_purpose
                                modified = True
                            
                            # Случай 3: Заполнены все три списка
                            elif contractor and payment_purpose and contractor_match and purpose_match:
                                record['Внутренние назначения'] = internal_purpose
                                modified = True
                    
                    # Сохраняем изменения, только если были модификации
                    if modified:
                        table.data_json = data_json
                        table.save()
                
                messages.success(request, "Фильтр успешно создан и применен")
            except Exception as e:
                messages.error(request, f"Ошибка при применении фильтров: {str(e)}")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                field_name = form.fields[field].label or field
                error_messages.append(f"{field_name}: {', '.join(errors)}")
            messages.error(request, f"Форма заполнена неверно: {'; '.join(error_messages)}")
    return redirect('index')

def generate_report(request):
    if request.method == 'POST':
        form = ReportFilterForm(request.POST)
        if form.is_valid():
            try:
                # Получаем все данные
                all_data = []
                for table in TableData.objects.all():
                    all_data.extend(table.data_json)
                
                # Применяем фильтры
                start_date = form.cleaned_data['start_date']
                end_date = form.cleaned_data['end_date']
                currency = form.cleaned_data['currency']
                
                filtered_data = []
                for record in all_data:
                    if currency and record['Валюта'] != currency:
                        continue
                        
                    date_str = record['Дата']
                    if date_str == "nan":
                        filtered_data.append(record)
                        continue
                        
                    try:
                        date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        if start_date and date < start_date:
                            continue
                        if end_date and date > end_date:
                            continue
                        filtered_data.append(record)
                    except ValueError:
                        continue
                
                # Создаем отчет
                report_data = generate_report_data(filtered_data)
                
                # Возвращаем JSON для отображения
                return JsonResponse({'success': True, 'data': report_data})
                
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid form'})

def generate_report_data(data):
    # Группировка данных по валютам, типам и назначениям
    report_data = {}
    months = set()
    
    for record in data:
        date_str = record['Дата']
        if date_str == 'nan':
            continue
            
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d')
            month_key = date.strftime('%m.%Y')
            months.add(month_key)
        except ValueError:
            continue
            
        currency = record.get('Валюта', 'RUB')
        if currency == 'nan':
            currency = 'RUB'
            
        purpose = record.get('Внутренние назначения', 'Не указано')
        if not purpose or purpose == 'nan':
            purpose = 'Не указано'
            
        if currency not in report_data:
            report_data[currency] = {'Credit': {}, 'Debit': {}}
            
        # Обработка кредита
        if record['Кредит'] != 'nan':
            if purpose not in report_data[currency]['Credit']:
                report_data[currency]['Credit'][purpose] = {}
            if month_key not in report_data[currency]['Credit'][purpose]:
                report_data[currency]['Credit'][purpose][month_key] = 0
            report_data[currency]['Credit'][purpose][month_key] += float(record['Кредит'])
            
        # Обработка дебета
        if record['Дебит'] != 'nan':
            if purpose not in report_data[currency]['Debit']:
                report_data[currency]['Debit'][purpose] = {}
            if month_key not in report_data[currency]['Debit'][purpose]:
                report_data[currency]['Debit'][purpose][month_key] = 0
            report_data[currency]['Debit'][purpose][month_key] += float(record['Дебит'])
    
    # Сортируем месяцы
    months = sorted(list(months), key=lambda x: datetime.strptime(x, '%m.%Y'))
    
    # Добавляем итоги для каждой валюты
    for currency in report_data:
        for operation_type in ['Credit', 'Debit']:
            if 'Итог' not in report_data[currency][operation_type]:
                report_data[currency][operation_type]['Итог'] = {}
            
            for month in months:
                total = sum(
                    purpose_data.get(month, 0)
                    for purpose, purpose_data in report_data[currency][operation_type].items()
                    if purpose != 'Итог'
                )
                report_data[currency][operation_type]['Итог'][month] = total
    
    return {
        'data': report_data,
        'months': months
    }

def download_report(request):
    if request.method == 'POST':
        try:
            report_data = json.loads(request.POST.get('report_data'))
            
            # Создаем DataFrame с мультииндексом
            columns = ['Категория', 'Тип'] + report_data['months']
            rows = []
            
            # Добавляем данные по каждой категории
            for category in report_data['categories']:
                # Строка для Дебета
                debit_row = [category, 'Дебет']
                for month in report_data['months']:
                    debit_row.append(report_data['data']['debit'][month].get(category, 0))
                rows.append(debit_row)
                
                # Строка для Кредита
                credit_row = [category, 'Кредит']
                for month in report_data['months']:
                    credit_row.append(report_data['data']['credit'][month].get(category, 0))
                rows.append(credit_row)
                
                # Строка для Итого
                total_row = [category, 'Итого']
                for month in report_data['months']:
                    total_row.append(report_data['data']['total'][month].get(category, 0))
                rows.append(total_row)
            
            # Добавляем итоговые строки
            for type_name in ['Дебет', 'Кредит', 'Итого']:
                total_row = ['ИТОГО', type_name]
                for month in report_data['months']:
                    total = 0
                    for category in report_data['categories']:
                        if type_name == 'Дебет':
                            total += report_data['data']['debit'][month].get(category, 0)
                        elif type_name == 'Кредит':
                            total += report_data['data']['credit'][month].get(category, 0)
                        else:
                            total += report_data['data']['total'][month].get(category, 0)
                    total_row.append(total)
                rows.append(total_row)
            
            # Создаем DataFrame
            df = pd.DataFrame(rows, columns=columns)
            
            # Создаем Excel файл
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=report.xlsx'
            
            # Записываем в Excel с форматированием
            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Отчет')
                
                # Получаем рабочий лист
                worksheet = writer.sheets['Отчет']
                
                # Форматируем числовые значения
                for row in worksheet.iter_rows(min_row=2):  # Пропускаем заголовок
                    for cell in row[2:]:  # Пропускаем колонки Категория и Тип
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                
                # Устанавливаем ширину колонок
                for column in worksheet.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            return response
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
            
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def filters_list(request):
    filters = Filter.objects.all().order_by('-id')
    return render(request, 'table_app/filters_list.html', {
        'filters': filters
    })

def delete_filter(request, filter_id):
    if request.method == 'POST':
        try:
            # Получаем фильтр
            filter_obj = Filter.objects.get(id=filter_id)
            
            # Получаем значения фильтра перед удалением
            contractor = filter_obj.contractor.strip()
            payment_purpose = filter_obj.payment_purpose.strip()
            internal_purpose = filter_obj.internal_purpose.strip()
            
            # Получаем все записи из базы данных
            table_data = TableData.objects.all()
            
            # Для каждой записи в базе данных
            for table in table_data:
                modified = False
                data_json = table.data_json
                
                for record in data_json:
                    # Проверяем условия фильтрации
                    contractor_match = contractor and record.get('Контрагент', '').strip() == contractor
                    purpose_match = payment_purpose and record.get('Банковское назначение платежа', '').strip() == payment_purpose
                    current_internal_purpose = record.get('Внутренние назначения', '').strip()
                    
                    # Если текущее внутреннее назначение совпадает с удаляемым фильтром
                    if current_internal_purpose == internal_purpose:
                        # Проверяем условия применения фильтра
                        if (contractor and not payment_purpose and contractor_match) or \
                           (payment_purpose and not contractor and purpose_match) or \
                           (contractor and payment_purpose and contractor_match and purpose_match):
                            # Сбрасываем значение внутреннего назначения
                            record['Внутренние назначения'] = ''
                            modified = True
                
                # Сохраняем изменения, только если были модификации
                if modified:
                    table.data_json = data_json
                    table.save()
            
            # Удаляем сам фильтр
            filter_obj.delete()
            messages.success(request, "Фильтр успешно удален, данные обновлены")
        except Filter.DoesNotExist:
            messages.error(request, "Фильтр не найден")
        except Exception as e:
            messages.error(request, f"Ошибка при удалении фильтра: {str(e)}")
    return redirect('filters_list')

def tables_list(request):
    tables = TableData.objects.all().order_by('-uploaded_at')
    
    # Добавляем информацию о полях для каждой таблицы
    for table in tables:
        if table.data_json:
            # Получаем все уникальные поля из данных
            fields = set()
            for record in table.data_json:
                fields.update(record.keys())
            
            # Определяем обязательные и опциональные поля
            required_fields = {"Дата", "Дебит", "Кредит", "Банковское назначение платежа"}
            table.fields = sorted(list(fields))
            table.required_fields = required_fields
    
    return render(request, 'table_app/tables_list.html', {
        'tables': tables
    })

def delete_table(request, table_id):
    if request.method == 'POST':
        try:
            table = TableData.objects.get(id=table_id)
            table.delete()
            messages.success(request, "Таблица успешно удалена")
        except TableData.DoesNotExist:
            messages.error(request, "Таблица не найдена")
        except Exception as e:
            messages.error(request, f"Ошибка при удалении таблицы: {str(e)}")
    return redirect('tables_list')

def append_to_table(request, table_id):
    if request.method == 'POST':
        try:
            table = TableData.objects.get(id=table_id)
            form = TableDataAppendForm(request.POST, request.FILES)
            
            if form.is_valid():
                file = request.FILES['file']
                file_extension = file.name.split('.')[-1].lower()
                
                if file_extension == 'csv':
                    df = pd.read_csv(file)
                elif file_extension == 'ods':
                    df = pd.read_excel(file, engine='odf')
                else:
                    df = pd.read_excel(file)
                
                # Добавление опциональных колонок
                optional_columns = ["Валюта", "Контрагент", "Внутренние назначения"]
                for col in optional_columns:
                    if col not in df.columns:
                        df[col] = ""
                
                # Преобразование в JSON
                new_records = []
                for _, row in df.iterrows():
                    record = {}
                    for col in df.columns:
                        value = row[col]
                        if isinstance(value, pd.Timestamp):
                            value = value.strftime('%Y-%m-%d')
                        elif pd.isna(value):
                            value = "nan"
                        else:
                            value = str(value)
                        record[col] = value
                    new_records.append(record)
                
                # Добавляем новые записи к существующим
                table.data_json.extend(new_records)
                table.save()
                
                messages.success(request, f"Добавлено {len(new_records)} записей в таблицу")
            else:
                error_messages = []
                for field, errors in form.errors.items():
                    field_name = form.fields[field].label or field
                    error_messages.append(f"{field_name}: {', '.join(errors)}")
                messages.error(request, f"Форма заполнена неверно: {'; '.join(error_messages)}")
                
        except TableData.DoesNotExist:
            messages.error(request, "Таблица не найдена")
        except Exception as e:
            messages.error(request, f"Ошибка при добавлении данных: {str(e)}")
            
    return redirect('tables_list') 